# -*- coding: utf-8 -*-
"""
Created on Sat May  8 20:29:59 2021

@author: 86155
"""

import openpyxl
from openpyxl.utils import get_column_letter

wb1 = openpyxl.load_workbook('附件1.xlsx')
ws1 = wb1.get_sheet_by_name('Sheet1')
wb2 = openpyxl.load_workbook('book3.xlsx')
ws2 = wb2.get_sheet_by_name('Sheet1')
print('start')

pname = []
dic = {}

ans = openpyxl.Workbook()
w = ans['Sheet']
ans.remove(w)

sheet1 = ans.create_sheet('采矿业')
sheet2 = ans.create_sheet('电力、热力、燃气及水生产和供应业')
sheet3 = ans.create_sheet('房地产业')
sheet4 = ans.create_sheet('建筑业')
sheet5 = ans.create_sheet('交通运输、仓储和邮政业')
sheet6 = ans.create_sheet('教育')
sheet7 = ans.create_sheet('金融业')
sheet8 = ans.create_sheet('居民服务、修理和其他服务业')
sheet9 = ans.create_sheet('科学研究和技术服务业')
sheet10 = ans.create_sheet('农、林、牧、渔业')
sheet11 = ans.create_sheet('批发和零售业')
sheet12 = ans.create_sheet('水利、环境和公共设施管理业')
sheet13 = ans.create_sheet('卫生和社会工作')
sheet14 = ans.create_sheet('文化、体育和娱乐业')
sheet15 = ans.create_sheet('信息传输、软件和信息技术服务业')
sheet16 = ans.create_sheet('制造业')
sheet17 = ans.create_sheet('住宿和餐饮业')
sheet18 = ans.create_sheet('综合')
sheet19 = ans.create_sheet('租赁和商务服务业')


for row in range(2, ws1.max_row + 1):
    pname.append(int(ws1['A' + str(row)].value + 0.1))
    dic[int(ws1['A' + str(row)].value + 0.1)] = ws1['B' + str(row)].value
print('start')

n = ws2.max_row + 1
for row in range(2, n):
    t = int(ws2['A' + str(row)].value + 0.1)
    if t in pname:
        if row % 100 == 0:
            print(row)
        data = []
        for i in range(1, ws2.max_column+1):
            data.append(ws2[str(get_column_letter(i)) + str(row)].value)
        if(dic[t] == '制造业'):
            sheet16.append(data)
        elif(dic[t] == '采矿业'):
            sheet1.append(data)
        elif(dic[t] == '电力、热力、燃气及水生产和供应业'):
            sheet2.append(data)
        elif(dic[t] == '房地产业'):
            sheet3.append(data)
        elif(dic[t] == '建筑业'):
            sheet4.append(data)
        elif(dic[t] == '交通运输、仓储和邮政业'):
            sheet5.append(data)
        elif(dic[t] == '教育'):
            sheet6.append(data)
        elif(dic[t] == '金融业'):
            sheet7.append(data)
        elif(dic[t] == '居民服务、修理和其他服务业'):
            sheet8.append(data)
        elif(dic[t] == '科学研究和技术服务业'):
            sheet9.append(data)
        elif(dic[t] == '农、林、牧、渔业'):
            sheet10.append(data)
        elif(dic[t] == '批发和零售业'):
            sheet11.append(data)
        elif(dic[t] == '水利、环境和公共设施管理业'):
            sheet12.append(data)
        elif(dic[t] == '卫生和社会工作'):
            sheet13.append(data)
        elif(dic[t] == '文化、体育和娱乐业'):
            sheet14.append(data)
        elif(dic[t] == '信息传输、软件和信息技术服务业'):
            sheet15.append(data)
        elif(dic[t] == '住宿和餐饮业'):
            sheet17.append(data)
        elif(dic[t] == '综合'):
            sheet18.append(data)
        elif(dic[t] == '租赁和商务服务业'):
            sheet19.append(data)

ans.save(filename='ans.xlsx')
print('end')