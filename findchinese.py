# -*- coding: utf-8 -*-
"""
Created on Sun May  9 20:34:20 2021

@author: 86155
"""

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('1.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
wb1 = openpyxl.load_workbook('附件3.xlsx')
ws1 = wb1.get_sheet_by_name('Sheet1')
print('start')

title = openpyxl.Workbook()
s = title.create_sheet('S')
data = []
dic = {}

for row in range(2, ws1.max_row + 1):
    dic[ws1['A' + str(row)].value] = ws1['B' + str(row)].value
for i in range(1, ws.max_column + 1):
    data.append(dic[ws[get_column_letter(i) + '1'].value])

print(data)
s.append(data)
print(dic)

title.save(filename='title.xlsx')